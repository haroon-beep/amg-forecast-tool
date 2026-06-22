from http.server import BaseHTTPRequestHandler
import json, re, base64, io
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

MONTH_NAMES = ['','January','February','March','April','May','June','July',
               'August','September','October','November','December']
MONTH_TEXT = {
    1:['JAN','JANUARY'],2:['FEB','FEBRUARY'],3:['MAR','MARCH'],4:['APR','APRIL'],
    5:['MAY'],6:['JUNE','JUN'],7:['JULY','JUL'],8:['AUG','AUGUST'],
    9:['SEPT','SEP','SEPTEMBER'],10:['OCT','OCTOBER'],
    11:['NOV','NOVEMBER'],12:['DEC','DECEMBER']
}
ACTUAL_LABELS  = ['ACTUAL ORDER SHIPPED','ACTUAL ORDER','ACTUAL SHIP']
FORECAST_LABELS = ['FORECAST']
ALL_LABELS = ACTUAL_LABELS + FORECAST_LABELS + ['BEG. IN STOCK','INCOMING','ON ORDER','IN STOCK']

def extract_amg(text):
    m = re.search(r'\b([A-Z]{2,4}\d{4,}(?:-CA\d|-TX)?)\b', str(text))
    return m.group(1) if m else None

def extract_nums(text):
    nums = re.findall(r'\b(\d{5,10})\b', str(text))
    return [n for n in nums if not (len(n)==4 and n[:2] in ('20','19'))]

def find_month_col(row_vals, month_num, year_num):
    aliases = [a.upper() for a in MONTH_TEXT.get(month_num, [])]
    for ci, v in enumerate(row_vals):
        if v is None: continue
        if isinstance(v, str):
            vu = v.strip().upper()
            if any(vu == a or vu.startswith(a) for a in aliases):
                return ci
        if isinstance(v, datetime):
            if v.month == month_num and v.year == year_num:
                return ci
    return None

def load_sap(data_bytes, month_num):
    wb = load_workbook(io.BytesIO(data_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(c).strip() if c else '' for c in rows[0]]
    mname = MONTH_NAMES[month_num].lower()
    qty_col = next((i for i,h in enumerate(headers)
                    if mname in h.lower() and 'quantity' in h.lower()), None)
    if qty_col is None:
        return {}, f"No quantity column for {MONTH_NAMES[month_num]}"
    data = {}
    for row in rows[1:]:
        item_no = str(row[1]).strip() if row[1] else None
        if item_no and item_no != 'None':
            qty = row[qty_col]
            if qty is not None:
                try: data[item_no] = float(qty)
                except: pass
    wb.close()
    return data, None

def process(sap_bytes, fc_bytes, tab_name, month_num, year_num):
    sap_data, err = load_sap(sap_bytes, month_num)
    if err:
        return None, err

    # Load with data_only=False to preserve ALL formulas and formatting
    wb = load_workbook(io.BytesIO(fc_bytes), data_only=False, keep_vba=False)

    sheet_name = next((n for n in wb.sheetnames
                       if tab_name.lower().strip() in n.lower().strip()), None)
    if not sheet_name:
        return None, f'Tab "{tab_name}" not found. Available: {", ".join(wb.sheetnames)}'

    ws = wb[sheet_name]

    # Read all values for matching (separate pass, read_only for speed)
    wb_ro = load_workbook(io.BytesIO(fc_bytes), read_only=True, data_only=True)
    ws_ro = wb_ro[sheet_name]
    all_rows = list(ws_ro.iter_rows(values_only=True))
    wb_ro.close()

    results = []
    updates  = []  # (excel_row_1based, excel_col_1based, value)

    i = 0
    while i < len(all_rows):
        row = all_rows[i]
        if not row or not row[0]: i += 1; continue
        cell0 = str(row[0]).strip()
        amg   = extract_amg(cell0)
        if not amg or any(cell0.upper().startswith(l) for l in ALL_LABELS):
            i += 1; continue

        month_col = find_month_col(row, month_num, year_num)
        if month_col is None: i += 1; continue

        item_nums = extract_nums(cell0)

        # Match SAP
        sap_qty = None
        for num in item_nums:
            if num in sap_data: sap_qty = sap_data[num]; break
        if sap_qty is None and amg in sap_data:
            sap_qty = sap_data[amg]

        # Find FORECAST and ACTUAL rows
        forecast_val   = None
        actual_row_idx = None
        for j in range(i+1, min(i+8, len(all_rows))):
            r = all_rows[j]
            if not r or not r[0]: continue
            lbl = str(r[0]).strip().upper()
            if extract_amg(str(r[0])) and not any(lbl.startswith(l) for l in ALL_LABELS):
                break
            raw = r[month_col] if len(r) > month_col else None
            try:
                val = float(raw) if raw is not None and str(raw).strip() not in ('','None','CUT FROM PO') else None
            except: val = None
            if any(lbl.startswith(l) for l in FORECAST_LABELS) and 'ACTUAL' not in lbl:
                forecast_val = val if val is not None else 0
            elif any(lbl.startswith(l) for l in ACTUAL_LABELS):
                actual_row_idx = j

        status = 'no_sap'
        variance = variance_pct = None
        if sap_qty is not None:
            variance = sap_qty - (forecast_val or 0)
            variance_pct = round((variance/forecast_val*100),1) if forecast_val else None
            status = 'over' if variance > 0 else ('under' if variance < 0 else 'match')

        results.append({
            'amg': amg,
            'item_num': item_nums[0] if item_nums else None,
            'description': cell0.replace('\n',' ').strip()[:65],
            'forecast': forecast_val,
            'actual': sap_qty,
            'variance': variance,
            'variance_pct': variance_pct,
            'status': status,
            'written': actual_row_idx is not None and sap_qty is not None
        })

        if actual_row_idx is not None and sap_qty is not None:
            # +1 for 1-based Excel rows
            updates.append((actual_row_idx + 1, month_col + 1, sap_qty))

        i += 1

    # ── WRITE ONLY THE ACTUAL VALUES — nothing else touched ──
    for (excel_row, excel_col, val) in updates:
        cell = ws.cell(row=excel_row, column=excel_col)
        cell.value = val
        # Preserve existing number format; only change value

    # Save to bytes
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    file_bytes = out.read()

    return {
        'file_b64': base64.b64encode(file_bytes).decode(),
        'sheet': sheet_name,
        'month': f'{MONTH_NAMES[month_num]} {year_num}',
        'items_found': len(results),
        'written': len(updates),
        'over': sum(1 for r in results if r['status']=='over'),
        'under': sum(1 for r in results if r['status']=='under'),
        'match': sum(1 for r in results if r['status']=='match'),
        'no_sap': sum(1 for r in results if r['status']=='no_sap'),
        'results': results
    }, None


class handler(BaseHTTPRequestHandler):

    def do_OPTIONS(self):
        self.send_response(200)
        self._cors()
        self.end_headers()

    def do_POST(self):
        length = int(self.headers.get('Content-Length', 0))
        body   = self.rfile.read(length)
        try:
            payload   = json.loads(body)
            sap_bytes = base64.b64decode(payload['sap_b64'])
            fc_bytes  = base64.b64decode(payload['fc_b64'])
            tab       = payload['tab']
            month_num = int(payload['month'])
            year_num  = int(payload['year'])

            result, err = process(sap_bytes, fc_bytes, tab, month_num, year_num)
            if err:
                self._respond(400, {'error': err})
            else:
                self._respond(200, result)
        except Exception as e:
            self._respond(500, {'error': str(e)})

    def _cors(self):
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')

    def _respond(self, code, data):
        body = json.dumps(data).encode()
        self.send_response(code)
        self._cors()
        self.send_header('Content-Type', 'application/json')
        self.send_header('Content-Length', str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, *args): pass
